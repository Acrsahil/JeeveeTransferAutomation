import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


def preview_output(path: str, rows: int = 5):
    """Read an Excel file, drop the 'median' column, add borders to all
    cells, auto-size each column to fit its content, save back to the
    same path, and print the first `rows` rows.
    """
    df = pd.read_excel(path)

    # ── Drop median column if present ──────────────────────────────────
    if 'median' in df.columns:
        df = df.drop(columns=['median'])

    # ── Save cleaned data back to the same file ────────────────────────
    df.to_excel(path, index=False)

    # ── Apply borders + auto column widths ─────────────────────────────
    wb = load_workbook(path)
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # Auto-fit each column width to its longest cell value
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        # padding for readability, capped so one giant cell can't blow up the sheet
        ws.column_dimensions[col_letter].width = min(max_len + 1, 60)

    wb.save(path)

    print(df.head(rows))
    return df


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python edit_outputfile.py <path_to_xlsx>")
        sys.exit(1)

    preview_output(sys.argv[1])
